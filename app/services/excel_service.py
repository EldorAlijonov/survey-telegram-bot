from datetime import date
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.user_repository import UserRepository
from app.utils.date_utils import format_dt
from app.utils.text import answer_label


class ExcelService:
    HEADERS = [
        "№",
        "Telegram ID",
        "Username",
        "Ism familya",
        "Telefon",
        "Zamonaviy kasblarga qiziqish",
        "Qulay vaqt",
        "Qiziqqan yo'nalish",
        "Obuna holati",
        "So'rovnoma topshirilgan sana",
    ]

    def __init__(self, session: AsyncSession) -> None:
        self.users = UserRepository(session)

    async def export_users(self, *, day: date | None = None, completed_only: bool = False) -> Path:
        users = await self.users.list_for_export(day=day, completed_only=completed_only)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Survey"
        sheet.append(self.HEADERS)

        for index, user in enumerate(users, start=1):
            survey = user.surveys[-1] if user.surveys else None
            sheet.append(
                [
                    index,
                    user.telegram_id,
                    f"@{user.username}" if user.username else "-",
                    user.full_name or "-",
                    user.phone or "-",
                    answer_label(survey.interest_level if survey else None),
                    answer_label(survey.convenient_time if survey else None),
                    answer_label(survey.interested_field if survey else None),
                    "Obuna bo'lgan" if user.is_subscribed else "Obuna bo'lmagan",
                    format_dt(survey.submitted_at if survey else None),
                ]
            )

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 42)
        sheet.freeze_panes = "A2"

        tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()
        path = Path(tmp.name)
        workbook.save(path)
        return path
